import xlwt
import random
import requests
import openpyxl
import multiprocessing
from math import ceil
from lxml import etree
from tenacity import retry, wait_fixed, stop_after_attempt
from monilogin import get_cookie

user_agent_list = [
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0;) Gecko/20100101 Firefox/61.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.62 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36',
    "Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10.5; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15",
]

USERNAME = '13620226479'
PASSWORD = '123456'


def get_proxy():
    url = 'http://h.shanchendaili.com/api.html?action=get_ip&key=HU9715b33b0206185636lEh5&time=10&count=1&protocol=http&type=json&only=0'
    response = requests.get(url).json()
    ip = response['list'][0]['sever']
    port = response['list'][0]['port']
    ip_1 = str(ip) + ':' + str(port)
    prox_ = f'{USERNAME}:{PASSWORD}@{ip_1}'
    proxy = {
        'http': 'http://' + prox_,
        'https': 'https://' + prox_
    }
    return proxy


class QiChaMaoSpider:
    def __init__(self):
        self.headers = ''
        self.company = ''
        self.holder_list = []
        self.invest_list = []
        self.cookies = get_cookie()

    @staticmethod
    def save_excel(datalist, sheet_name):
        file = xlwt.Workbook(encoding='utf-8')
        sheet = file.add_sheet(sheet_name)
        j = 0
        for i in dict(datalist[0]):  # 写入列名
            sheet.write(0, j, i)
            j += 1

        ro = 0
        for dat in datalist:
            ro += 1
            co = 0
            for val in dat.values():
                sheet.write(ro, co, str(val))
                co += 1
        file.save(sheet_name + '.xls')

    def parse_detail(self, url):
        try:
            page = etree.HTML(self.download_page(url))
            holders = page.xpath("//div[@id='M_gdxx']//div[@class='data-list']/ul")  # 股东信息
            for holder in holders:
                em = holder.xpath("./li[2]//em/text()")[0]  # 股东类型
                if em == '自然法人' or em == '自然人股东':
                    continue
                else:
                    hold = holder.xpath("./li[2]//a/text()")[0]  # 股东
                    _url = holder.xpath("./li[2]//a/@href")[0]
                    url = 'https://www.qichamao.com' + _url
                    industry, address = self.get_ind_add(url)
                    data = {
                        '股东公司（入）': hold,
                        '城市（地址）': address,
                        '所属行业': ''.join(industry).strip()
                    }
                    print(data)
                    self.holder_list.append(data)

            dwtz_num = page.xpath("//div[@id='M_dwtz']//h2/em/text()")  # 对外投资数
            if dwtz_num:
                investers = page.xpath("//div[@id='M_dwtz']//div[@class='data-list']/ul")
                for invest in investers:
                    inves = invest.xpath("./li[2]//a/text()")[0]  # 投资人
                    _url = invest.xpath("./li[2]//a/@href")[0]
                    url = 'https://www.qichamao.com' + _url
                    industry, address = self.get_ind_add(url)
                    data = {
                        '投资公司（出）': inves,
                        '城市（地址）': address,
                        '所属行业': ''.join(industry).strip()
                    }
                    print(data)
                    self.invest_list.append(data)
                if int(dwtz_num[0]) > 10:
                    self.parse_invest(dwtz_num[0])  # 对外投资公司，数量大于10需另外请求接口
        except Exception as e:
            print(e)

    def get_ind_add(self, url):  # 获取地址跟所属行业
        try:
            page = etree.HTML(self.download_page(url))
            industry = page.xpath("//div[@id='tagContent']/div[1]/ul/li[1]/div//text()")  # 所属行业
            address = page.xpath("//li[@class='w-all'][1]/div//text()")[0]  # 地址
            return industry, address
        except Exception as e:
            print(e)

    def parse_invest(self, num):
        try:
            page_num = ceil(int(num) / 10)  # 页数，一页10条数据，向上取整
            for i in range(2, page_num + 1):
                form_data = {
                    'PageNo': i,
                    'oc_name': self.company,
                    'currpage': i,
                    'pagesize': 10
                }
                url = 'https://www.qichamao.com/orgcompany/GetDWTZInfo'
                proxies = get_proxy()
                res = requests.post(url, data=form_data, headers=self.headers, proxies=proxies).json()
                companyList = res['dataList']['CompanyList']
                for comp in companyList:
                    name = comp['oc_name']
                    code = comp['oc_code']
                    url = f'https://www.qichamao.com/orgcompany/searchitemdtl/{code}.html'
                    industry, address = self.get_ind_add(url)
                    data = {
                        '投资公司（出）': name,
                        '城市（地址）': address,
                        '所属行业': ''.join(industry).strip()
                    }
                    print(data)
                    self.invest_list.append(data)
        except Exception as e:
            print(e)

    @retry(stop=stop_after_attempt(10), wait=wait_fixed(3))
    def download_page(self, url):
        try:
            proxies = get_proxy()
            self.headers = {'user-agent': random.choice(user_agent_list)}
            response = requests.get(url, headers=self.headers, cookies=self.cookies, proxies=proxies)
            if '登录-企查猫(企业查询宝)' not in response.text and '用户验证-企查猫(企业查询宝)' not in response.text:
                return response.text
            else:
                self.cookies = get_cookie()
                raise Exception
        except Exception as e:
            print(e)
            raise e

    def parse_list(self):
        print('start!')
        try:
            wb = openpyxl.load_workbook("test01.xlsx")
            ws = wb.active
            for r in range(2, ws.max_row):
                self.company = ws.cell(r, 2).value
                print(self.company)
                list_url = f'https://www.qichamao.com/search/all/{self.company}'
                res_text = self.download_page(list_url)
                response_page = etree.HTML(res_text)
                _url = response_page.xpath("//ul[@id='listsec']/li[1]//a[@class='listsec_tit']/@href")[0]
                content_url = 'https://www.qichamao.com' + _url
                self.parse_detail(content_url)

            self.save_excel(self.holder_list, '股东信息')
            self.save_excel(self.invest_list, '对外投资')
        except Exception as e:
            print(e)
        print('finish!')

    def start(self):
        process = multiprocessing.Process(target=self.parse_list)
        process.start()


if __name__ == '__main__':
    qichamao = QiChaMaoSpider()
    # qichamao.parse_list()
    qichamao.start()
